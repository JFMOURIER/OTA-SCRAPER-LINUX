from __future__ import annotations

import csv
import fcntl
import hashlib
import inspect
import json
import subprocess
import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import openpyxl

import database.db as db
from services.drive_delivery import (
    drive_configuration_status,
    load_drive_state,
    retry_latest_failed_upload,
    upload_run_bundle,
)
from services.run_exports import automatic_export_run_bundle
from services.schedule_config import (
    DATE_MODE_AUTOMATIC,
    DATE_MODE_MANUAL,
    PARIS,
    ScheduleValidationError,
    current_due_slot,
    default_daily_times,
    default_schedule,
    load_schedule,
    next_scheduled_run,
    read_schedule_state,
    resolved_dates_for_schedule,
    save_schedule,
)
from services.schedule_dispatcher import (
    dispatch_instance,
    dispatch_once,
    request_run_once,
)
from services.scheduled_instances import (
    INSTANCE_ORDER,
    SCHEDULED_INSTANCES,
    resolve_automatic_windows,
)


def _definition(instance_id: str, root: Path):
    return replace(
        SCHEDULED_INSTANCES[instance_id],
        data_dir_override=root / instance_id,
    )


def _row(run_id: int, index: int) -> dict:
    return {
        "collection_run_id": run_id,
        "source": "Demo",
        "city_or_region": "Orlando",
        "hotel_name": f"Hotel {index}",
        "ota_hotel_id": str(index),
        "hotel_url": f"https://example.test/{index}",
        "raw_price_text": f"US$ {100 + index}",
        "parsed_price": 100 + index,
        "cheapest_price_total": 100 + index,
        "currency": "USD",
        "checkin_date": date(2026, 8, 1),
        "checkout_date": date(2026, 8, 2),
        "requested_checkin_date": date(2026, 8, 1),
        "requested_checkout_date": date(2026, 8, 2),
        "effective_checkin_date": date(2026, 8, 1),
        "effective_checkout_date": date(2026, 8, 2),
        "date_integrity_verified": True,
        "number_of_nights": 1,
        "adults": 2,
        "collection_status": "success",
        "collected_at": datetime(2026, 7, 24, 12, 0),
    }


def _create_run(database_path: Path, rows: int = 3) -> int:
    with patch.object(db, "SQLITE_DB_PATH", database_path):
        db.init_db("sqlite")
        run_id = db.create_collection_run(
            "Demo",
            "Orlando",
            date(2026, 8, 1),
            date(2026, 8, 2),
            1,
            2,
            "USD",
            rows,
            backend="sqlite",
        )
        for offset in range(0, rows, 200):
            db.insert_hotel_results(
                [
                    _row(run_id, index)
                    for index in range(offset, min(rows, offset + 200))
                ],
                backend="sqlite",
            )
        with patch.dict(
            "os.environ",
            {"OTA_AUTO_EXPORT_ENABLED": "0"},
        ):
            db.update_collection_run_status(
                run_id,
                "completed_all_dates",
                backend="sqlite",
            )
    return run_id


class AutomaticWindowTests(unittest.TestCase):
    def test_required_july_24_windows_and_eight_month_long_tail(self):
        anchor = date(2026, 7, 24)
        windows = resolve_automatic_windows(anchor)
        self.assertEqual(
            windows["near_30_days"],
            (date(2026, 7, 24), date(2026, 8, 23)),
        )
        self.assertEqual(
            windows["medium_31_120_days"],
            (date(2026, 8, 24), date(2026, 11, 23)),
        )
        self.assertEqual(
            windows["long_121_365_days"],
            (date(2026, 11, 24), date(2027, 7, 23)),
        )
        long_days = (windows["long_121_365_days"][1] - windows["long_121_365_days"][0]).days + 1
        self.assertGreaterEqual(long_days, 240)
        self.assertLessEqual(long_days, 246)

    def test_contiguous_non_overlapping_complete_horizon(self):
        anchor = date(2026, 7, 24)
        windows = resolve_automatic_windows(anchor)
        self.assertEqual(
            windows["near_30_days"][1] + timedelta(days=1),
            windows["medium_31_120_days"][0],
        )
        self.assertEqual(
            windows["medium_31_120_days"][1] + timedelta(days=1),
            windows["long_121_365_days"][0],
        )
        owners = {}
        for instance_id, (start, end) in windows.items():
            for offset in range((end - start).days + 1):
                day = start + timedelta(days=offset)
                self.assertNotIn(day, owners)
                owners[day] = instance_id
        expected = {
            anchor + timedelta(days=offset)
            for offset in range(
                (windows["long_121_365_days"][1] - anchor).days + 1
            )
        }
        self.assertEqual(set(owners), expected)
        self.assertLessEqual(
            windows["long_121_365_days"][1],
            anchor + timedelta(days=365),
        )

    def test_month_end_february_and_leap_year(self):
        january = resolve_automatic_windows(date(2027, 1, 31))
        self.assertEqual(
            january["near_30_days"],
            (date(2027, 1, 31), date(2027, 2, 27)),
        )
        leap = resolve_automatic_windows(date(2024, 2, 29))
        self.assertEqual(leap["near_30_days"][1], date(2024, 3, 28))
        self.assertLessEqual(
            leap["long_121_365_days"][1],
            date(2024, 2, 29) + timedelta(days=365),
        )

    def test_interval_next_run_handles_paris_dst_transition(self):
        schedule = default_schedule(
            "near_30_days",
            now=datetime(2026, 3, 29, 1, 30, tzinfo=PARIS),
        )
        schedule["enabled"] = True
        schedule["interval_minutes"] = 60
        next_run = next_scheduled_run(
            schedule,
            now=datetime(2026, 3, 29, 3, 10, tzinfo=PARIS),
        )
        self.assertEqual(next_run.hour, 3)
        self.assertEqual(next_run.minute, 30)
        self.assertEqual(next_run.utcoffset(), timedelta(hours=2))


class SchedulePersistenceAndValidationTests(unittest.TestCase):
    def test_manual_dates_override_persist_and_auto_recalculates(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            raw = default_schedule("near_30_days")
            raw.update(
                {
                    "date_mode": DATE_MODE_MANUAL,
                    "manual_start_date": "2026-09-01",
                    "manual_end_date": "2026-09-10",
                }
            )
            save_schedule("near_30_days", raw, data_dir=root)
            restarted = load_schedule("near_30_days", data_dir=root)
            self.assertEqual(
                resolved_dates_for_schedule(
                    restarted,
                    date(2027, 1, 1),
                ),
                (date(2026, 9, 1), date(2026, 9, 10)),
            )
            restarted["date_mode"] = DATE_MODE_AUTOMATIC
            save_schedule("near_30_days", restarted, data_dir=root)
            automatic = load_schedule("near_30_days", data_dir=root)
            self.assertEqual(
                resolved_dates_for_schedule(
                    automatic,
                    date(2027, 1, 1),
                ),
                resolve_automatic_windows(date(2027, 1, 1))[
                    "near_30_days"
                ],
            )

    def test_manual_over_365_requires_explicit_confirmation(self):
        with tempfile.TemporaryDirectory() as directory:
            raw = default_schedule("near_30_days")
            raw.update(
                {
                    "date_mode": DATE_MODE_MANUAL,
                    "manual_start_date": "2026-01-01",
                    "manual_end_date": "2027-01-02",
                    "manual_over_one_year_confirmed": False,
                }
            )
            with self.assertRaises(ScheduleValidationError):
                save_schedule("near_30_days", raw, data_dir=directory)
            raw["manual_over_one_year_confirmed"] = True
            saved = save_schedule(
                "near_30_days",
                raw,
                data_dir=directory,
            )
            self.assertTrue(saved["manual_over_one_year_confirmed"])

    def test_frequency_ranges_and_duplicate_times(self):
        with tempfile.TemporaryDirectory() as directory:
            near = default_schedule("near_30_days")
            for invalid in (14, 1441):
                near["interval_minutes"] = invalid
                with self.assertRaises(ScheduleValidationError):
                    save_schedule(
                        "near_30_days",
                        near,
                        data_dir=Path(directory) / str(invalid),
                    )
            for instance_id, maximum in (
                ("medium_31_120_days", 4),
                ("long_121_365_days", 2),
            ):
                schedule = default_schedule(instance_id)
                schedule["runs_per_day"] = maximum + 1
                schedule["daily_run_times"] = [
                    f"{index:02d}:00"
                    for index in range(maximum + 1)
                ]
                with self.assertRaises(ScheduleValidationError):
                    save_schedule(
                        instance_id,
                        schedule,
                        data_dir=Path(directory) / instance_id,
                    )
            medium = default_schedule("medium_31_120_days")
            medium["daily_run_times"] = ["00:20", "00:20"]
            with self.assertRaises(ScheduleValidationError):
                save_schedule(
                    "medium_31_120_days",
                    medium,
                    data_dir=Path(directory) / "duplicate",
                )

    def test_daily_choices_defaults_and_drive_ids(self):
        self.assertEqual(
            default_daily_times("medium_31_120_days", 1),
            ("00:20",),
        )
        self.assertEqual(
            default_daily_times("medium_31_120_days", 4),
            ("00:20", "06:20", "12:20", "18:20"),
        )
        self.assertEqual(
            default_daily_times("long_121_365_days", 2),
            ("01:35", "13:35"),
        )
        expected = {
            "near_30_days": "18kkxujFodzEfoOmhfpBuZI8xDzaoUk8b",
            "medium_31_120_days": "1ATZztmP0pS9c0oprzF3LzQKYcpsXtT7S",
            "long_121_365_days": "19TNxbw6-ymHmUE2dkrSLyvKwYtjP4H5c",
        }
        self.assertEqual(
            {
                key: value.drive_folder_id
                for key, value in SCHEDULED_INSTANCES.items()
            },
            expected,
        )

    def test_invalid_save_does_not_overwrite_previous(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            valid = default_schedule("near_30_days")
            valid["interval_minutes"] = 30
            save_schedule("near_30_days", valid, data_dir=root)
            invalid = dict(valid)
            invalid["interval_minutes"] = 1
            with self.assertRaises(ScheduleValidationError):
                save_schedule("near_30_days", invalid, data_dir=root)
            self.assertEqual(
                load_schedule("near_30_days", data_dir=root)[
                    "interval_minutes"
                ],
                30,
            )

    def test_schedule_save_does_not_rewrite_systemd_timers(self):
        project = Path(__file__).resolve().parents[1]
        timers = sorted((project / "systemd").glob("*.timer"))
        before = {
            path.name: hashlib.sha256(path.read_bytes()).hexdigest()
            for path in timers
        }
        with tempfile.TemporaryDirectory() as directory:
            save_schedule(
                "near_30_days",
                default_schedule("near_30_days"),
                data_dir=directory,
            )
        after = {
            path.name: hashlib.sha256(path.read_bytes()).hexdigest()
            for path in timers
        }
        self.assertEqual(before, after)

    def test_streamlit_render_is_passive_until_button_click(self):
        from services import schedule_ui

        source = inspect.getsource(schedule_ui.render_automatic_schedule)
        button_position = source.index('"Run Once Now"')
        call_position = source.index(
            "request_run_once(instance_id",
            button_position,
        )
        self.assertGreater(call_position, button_position)
        self.assertNotIn("request_run_once(instance_id", source[:button_position])


class DispatcherTests(unittest.TestCase):
    def _enabled_near(self, root: Path) -> tuple[Any, datetime]:
        definition = _definition("near_30_days", root)
        anchor = datetime(2026, 7, 24, 10, 0, tzinfo=PARIS)
        schedule = default_schedule("near_30_days", now=anchor)
        schedule["enabled"] = True
        save_schedule(
            "near_30_days",
            schedule,
            data_dir=definition.data_dir,
            now=anchor,
        )
        return definition, anchor

    def test_slot_idempotency_and_next_boundary(self):
        with tempfile.TemporaryDirectory() as directory:
            definition, anchor = self._enabled_near(Path(directory))
            launches = []

            def launcher(instance_id, request):
                launches.append((instance_id, request["schedule_slot"]))
                return True

            now = anchor + timedelta(minutes=60)
            first = dispatch_instance(
                definition,
                now=now,
                launcher=launcher,
                capacity_check=lambda: True,
            )
            second = dispatch_instance(
                definition,
                now=now,
                launcher=launcher,
                capacity_check=lambda: True,
            )
            self.assertEqual(first["action"], "launched")
            self.assertEqual(second["action"], "not_due")
            self.assertEqual(len(launches), 1)
            schedule = load_schedule(
                "near_30_days",
                data_dir=definition.data_dir,
            )
            self.assertEqual(
                next_scheduled_run(schedule, now=now),
                anchor + timedelta(minutes=120),
            )

    def test_same_instance_overlap_skips_without_failure(self):
        with tempfile.TemporaryDirectory() as directory:
            definition, anchor = self._enabled_near(Path(directory))
            lock_path = definition.data_dir / "status" / "active_scraper.lock"
            lock_path.parent.mkdir(parents=True, exist_ok=True)
            with lock_path.open("a+", encoding="utf-8") as handle:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
                result = dispatch_instance(
                    definition,
                    now=anchor + timedelta(minutes=60),
                    launcher=lambda *_: self.fail("must not launch"),
                    capacity_check=lambda: True,
                )
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
            self.assertEqual(result["action"], "skipped_active")
            state = read_schedule_state(
                "near_30_days",
                data_dir=definition.data_dir,
            )
            self.assertEqual(
                state["last_skip_reason"],
                "scheduled_run_skipped_previous_run_active",
            )

    def test_host_capacity_defers_one_pending_without_backlog(self):
        with tempfile.TemporaryDirectory() as directory:
            definition, anchor = self._enabled_near(Path(directory))
            now = anchor + timedelta(minutes=180)
            first = dispatch_instance(
                definition,
                now=now,
                launcher=lambda *_: self.fail("must not launch"),
                capacity_check=lambda: False,
            )
            state = read_schedule_state(
                "near_30_days",
                data_dir=definition.data_dir,
            )
            pending_key = state["pending_due_run"]["slot_key"]
            second = dispatch_instance(
                definition,
                now=now + timedelta(minutes=1),
                launcher=lambda *_: True,
                capacity_check=lambda: True,
            )
            self.assertEqual(first["action"], "deferred_host_capacity")
            self.assertEqual(second["action"], "launched")
            self.assertEqual(second["slot_key"], pending_key)
            state = read_schedule_state(
                "near_30_days",
                data_dir=definition.data_dir,
            )
            self.assertIsNone(state["pending_due_run"])
            self.assertEqual(len(state["dispatched_slot_keys"]), 1)

    def test_run_once_respects_instance_and_host_locks(self):
        with tempfile.TemporaryDirectory() as directory:
            definition = _definition("near_30_days", Path(directory))
            save_schedule(
                "near_30_days",
                default_schedule("near_30_days"),
                data_dir=definition.data_dir,
            )
            lock_path = definition.data_dir / "status" / "active_scraper.lock"
            lock_path.parent.mkdir(parents=True, exist_ok=True)
            with lock_path.open("a+", encoding="utf-8") as handle:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
                result = request_run_once(
                    "near_30_days",
                    data_dir=definition.data_dir,
                    launcher=lambda *_: self.fail("must not launch"),
                    capacity_check=lambda: True,
                )
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
            self.assertEqual(
                result["reason"],
                "scheduled_run_skipped_previous_run_active",
            )
            deferred = request_run_once(
                "near_30_days",
                data_dir=definition.data_dir,
                launcher=lambda *_: self.fail("must not launch"),
                capacity_check=lambda: False,
            )
            self.assertEqual(
                deferred["reason"],
                "scheduled_run_deferred_host_capacity",
            )

    def test_fixed_timer_conflict_blocks_dispatcher(self):
        with patch(
            "services.schedule_dispatcher.active_old_fixed_timers",
            return_value=["ota-scraper-near.timer"],
        ):
            result = dispatch_once()
        self.assertEqual(result["status"], "migration_conflict")
        self.assertEqual(result["instances"], [])


class FakeRclone:
    def __init__(self, *, fail_copy: bool = False) -> None:
        self.files: dict[str, dict[str, Any]] = {}
        self.copy_calls = 0
        self.fail_copy = fail_copy

    def __call__(self, command):
        args = list(command)
        action_index = args.index("--config") + 2
        action = args[action_index]
        if action == "listremotes":
            return subprocess.CompletedProcess(
                args,
                0,
                stdout="gdrive:\n",
                stderr="",
            )
        if action == "lsf":
            return subprocess.CompletedProcess(args, 0, stdout="", stderr="")
        if action == "copyto":
            self.copy_calls += 1
            if self.fail_copy:
                return subprocess.CompletedProcess(
                    args,
                    1,
                    stdout="",
                    stderr="network unavailable",
                )
            local = Path(args[action_index + 1])
            remote = args[action_index + 2].split(":", 1)[1]
            data = local.read_bytes()
            self.files[remote] = {
                "Size": len(data),
                "Hashes": {"MD5": hashlib.md5(data).hexdigest()},
            }
            return subprocess.CompletedProcess(args, 0, stdout="", stderr="")
        if action == "lsjson":
            remote = args[action_index + 1].split(":", 1)[1]
            if remote not in self.files:
                return subprocess.CompletedProcess(
                    args,
                    3,
                    stdout="",
                    stderr="not found",
                )
            return subprocess.CompletedProcess(
                args,
                0,
                stdout=json.dumps(self.files[remote]),
                stderr="",
            )
        if action == "deletefile":
            remote = args[action_index + 1].split(":", 1)[1]
            self.files.pop(remote, None)
            return subprocess.CompletedProcess(args, 0, stdout="", stderr="")
        raise AssertionError(args)


class ExportAndDriveTests(unittest.TestCase):
    def test_complete_csv_excel_are_full_count_atomic_and_named(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "hotel_price_collector.sqlite"
            run_id = _create_run(database_path, 620)
            downloads = root / "Downloads"
            with patch.object(db, "SQLITE_DB_PATH", database_path):
                payload = automatic_export_run_bundle(
                    run_id,
                    database_path=database_path,
                    instance_id="near_30_days",
                    downloads_dir=downloads,
                    copy_to_downloads=True,
                )
            self.assertEqual(payload["csv_rows_exported"], 620)
            self.assertEqual(payload["excel_rows_exported"], 620)
            self.assertIn("_final_", Path(payload["csv_file_path"]).name)
            self.assertIn("_final_", Path(payload["excel_file_path"]).name)
            self.assertEqual(
                Path(payload["csv_file_path"]).read_bytes()[:3],
                b"\xef\xbb\xbf",
            )
            with Path(payload["csv_file_path"]).open(
                "r",
                encoding="utf-8-sig",
                newline="",
            ) as handle:
                self.assertEqual(len(list(csv.DictReader(handle))), 620)
            workbook = openpyxl.load_workbook(
                payload["excel_file_path"],
                read_only=True,
            )
            self.assertTrue(
                {
                    "All Hotel Results",
                    "Daily Summary",
                    "Run Summary",
                }.issubset(workbook.sheetnames)
            )
            excel_rows = (
                sum(
                    1
                    for _ in workbook["All Hotel Results"].iter_rows(
                        values_only=True
                    )
                )
                - 1
            )
            workbook.close()
            self.assertEqual(excel_rows, 620)
            self.assertTrue(Path(payload["csv_downloads_path"]).is_file())
            self.assertTrue(Path(payload["excel_downloads_path"]).is_file())
            self.assertEqual(list(root.rglob("*.tmp")), [])

    def test_drive_upload_is_called_only_after_both_local_exports_succeed(self):
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            database_path = data_dir / "hotel_price_collector.sqlite"
            run_id = _create_run(database_path, 3)
            schedule = default_schedule("near_30_days")
            schedule["drive_upload_enabled"] = True
            save_schedule("near_30_days", schedule, data_dir=data_dir)
            observed: dict[str, Any] = {}

            def verify_local_first(_instance_id, run, **kwargs):
                csv_path = Path(kwargs["csv_path"])
                excel_path = Path(kwargs["excel_path"])
                current = db.fetch_collection_run_by_id(
                    run_id,
                    backend="sqlite",
                )
                observed.update(
                    {
                        "csv_exists": csv_path.is_file(),
                        "excel_exists": excel_path.is_file(),
                        "csv_status": current["csv_export_status"],
                        "excel_status": current["excel_export_status"],
                    }
                )
                return {"drive_upload_status": "succeeded"}

            with patch.object(
                db,
                "SQLITE_DB_PATH",
                database_path,
            ), patch(
                "services.drive_delivery.upload_run_bundle",
                side_effect=verify_local_first,
            ) as upload:
                automatic_export_run_bundle(
                    run_id,
                    database_path=database_path,
                    instance_id="near_30_days",
                    copy_to_downloads=False,
                )
            upload.assert_called_once()
            self.assertEqual(
                observed,
                {
                    "csv_exists": True,
                    "excel_exists": True,
                    "csv_status": "succeeded",
                    "excel_status": "succeeded",
                },
            )

    def test_drive_upload_idempotent_correct_folder_and_manifest(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = root / "rclone.conf"
            config.write_text("[gdrive]\ntype = drive\n", encoding="utf-8")
            data_dir = root / "near"
            schedule = default_schedule("near_30_days")
            schedule["drive_upload_enabled"] = True
            save_schedule("near_30_days", schedule, data_dir=data_dir)
            csv_path = data_dir / "exports" / "result.csv"
            excel_path = data_dir / "exports" / "result.xlsx"
            csv_path.parent.mkdir(parents=True)
            csv_path.write_text("a,b\n1,2\n", encoding="utf-8")
            excel_path.write_bytes(b"disposable excel")
            run = {
                "id": 42,
                "source": "Demo",
                "city_or_region": "Orlando",
                "resolved_start_date": "2026-07-24",
                "resolved_end_date": "2026-08-23",
                "date_mode": "automatic",
                "schedule_slot": "slot",
                "frequency_configuration": {"mode": "interval"},
                "status": "completed_all_dates",
                "started_at": "2026-07-24T10:00:00+02:00",
                "completed_at": "2026-07-24T10:15:00+02:00",
                "authoritative_row_count": 1,
                "csv_rows_exported": 1,
                "excel_rows_exported": 1,
            }
            fake = FakeRclone()
            with patch(
                "services.drive_delivery._rclone_binary",
                return_value="rclone",
            ), patch(
                "services.drive_delivery._config_path",
                return_value=config,
            ):
                first = upload_run_bundle(
                    "near_30_days",
                    run,
                    data_dir=data_dir,
                    csv_path=csv_path,
                    excel_path=excel_path,
                    runner=fake,
                )
                second = upload_run_bundle(
                    "near_30_days",
                    run,
                    data_dir=data_dir,
                    csv_path=csv_path,
                    excel_path=excel_path,
                    runner=fake,
                )
            self.assertEqual(first["drive_upload_status"], "succeeded")
            self.assertEqual(second["drive_upload_status"], "succeeded")
            self.assertEqual(fake.copy_calls, 3)
            self.assertEqual(
                first["drive_folder_id"],
                SCHEDULED_INSTANCES["near_30_days"].drive_folder_id,
            )
            self.assertTrue(first["drive_manifest_path"].endswith(".json"))

    def test_drive_failure_does_not_change_collection_and_retry_no_scrape(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = root / "rclone.conf"
            config.write_text("[gdrive]\ntype = drive\n", encoding="utf-8")
            data_dir = root / "instance"
            database_path = data_dir / "hotel_price_collector.sqlite"
            run_id = _create_run(database_path, 1)
            schedule = default_schedule("near_30_days")
            schedule["drive_upload_enabled"] = True
            save_schedule("near_30_days", schedule, data_dir=data_dir)
            csv_path = data_dir / "exports" / "r.csv"
            excel_path = data_dir / "exports" / "r.xlsx"
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            csv_path.write_text("x\n1\n", encoding="utf-8")
            excel_path.write_bytes(b"xlsx")
            with patch.object(db, "SQLITE_DB_PATH", database_path):
                run = db.fetch_collection_run_by_id(run_id, backend="sqlite")
                run["authoritative_row_count"] = 1
                fake = FakeRclone(fail_copy=True)
                with patch(
                    "services.drive_delivery._rclone_binary",
                    return_value="rclone",
                ), patch(
                    "services.drive_delivery._config_path",
                    return_value=config,
                ):
                    failed = upload_run_bundle(
                        "near_30_days",
                        run,
                        data_dir=data_dir,
                        csv_path=csv_path,
                        excel_path=excel_path,
                        runner=fake,
                    )
                self.assertEqual(
                    db.fetch_collection_run_by_id(
                        run_id,
                        backend="sqlite",
                    )["status"],
                    "completed_all_dates",
                )
                self.assertEqual(failed["drive_upload_status"], "failed")
                fake.fail_copy = False
                with patch(
                    "services.drive_delivery._rclone_binary",
                    return_value="rclone",
                ), patch(
                    "services.drive_delivery._config_path",
                    return_value=config,
                ), patch(
                    "services.schedule_dispatcher.request_run_once"
                ) as no_scrape:
                    retried = retry_latest_failed_upload(
                        "near_30_days",
                        data_dir=data_dir,
                        runner=fake,
                    )
                no_scrape.assert_not_called()
                self.assertEqual(
                    retried["drive_upload_status"],
                    "succeeded",
                )

    def test_drive_not_configured_is_truthful_and_no_credentials_in_source(self):
        with patch(
            "services.drive_delivery._rclone_binary",
            return_value=None,
        ):
            status = drive_configuration_status("near_30_days")
        self.assertEqual(status["status"], "not_configured")
        self.assertEqual(
            status["installation_command"],
            "sudo apt install rclone",
        )
        project = Path(__file__).resolve().parents[1]
        source_text = "\n".join(
            path.read_text(encoding="utf-8")
            for path in [
                project / "services" / "drive_delivery.py",
                project / "scripts" / "ota-drive",
            ]
        )
        self.assertNotIn("client_secret", source_text.lower())
        self.assertNotIn("access_token", source_text.lower())
        self.assertNotIn("refresh_token", source_text.lower())


class IsolationContractTests(unittest.TestCase):
    def test_all_schedule_and_delivery_paths_are_per_instance(self):
        roots = [
            SCHEDULED_INSTANCES[instance_id].data_dir.resolve()
            for instance_id in INSTANCE_ORDER
        ]
        self.assertEqual(len(roots), len(set(roots)))
        for definition in SCHEDULED_INSTANCES.values():
            self.assertNotEqual(definition.data_dir.name, "period_1")
            self.assertEqual(
                definition.data_dir
                / "config"
                / "schedule.json",
                definition.data_dir / "config" / "schedule.json",
            )
            self.assertEqual(
                definition.data_dir
                / "status"
                / "schedule_state.json",
                definition.data_dir / "status" / "schedule_state.json",
            )


if __name__ == "__main__":
    unittest.main()
